
from __future__ import annotations

import os
import re
import copy
from io import BytesIO
from datetime import datetime
from typing import Dict, Optional, List
from urllib.parse import urljoin

import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_FILE = os.path.join(BASE_DIR, "매직_견적서_cs.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

SHEET_NAME = "매직 견적서"
FIRST_ITEM_ROW = 15
APP_PASSWORD = "0915"


def log(msg: str):
    st.session_state.setdefault("logs", [])
    st.session_state["logs"].append(msg)


def clean_price_to_int(text: Optional[str]) -> Optional[int]:
    if not text:
        return None
    digits = re.sub(r"[^0-9]", "", text)
    return int(digits) if digits else None


def vat_to_supply(vat_price: Optional[int]) -> Optional[int]:
    if vat_price is None:
        return None
    return int(round(vat_price / 1.1))


def clean_product_name(name: Optional[str]) -> Optional[str]:
    if not name:
        return name
    name = re.sub(r"\s*-\s*\(주\)엠퍼니처\s*$", "", name.strip())
    name = re.sub(r"\s+", " ", name)
    return name.strip()


def normalize_url(raw_url: Optional[str], base_url: str) -> Optional[str]:
    if not raw_url:
        return None
    raw_url = raw_url.strip()
    if raw_url.startswith("//"):
        return "https:" + raw_url
    return urljoin(base_url, raw_url)


def looks_like_size_line(text: str) -> bool:
    text = text.replace("×", "x")
    tokens = re.findall(r"(?:W|D|H|SH|AH|Ø)\s*\d+", text, flags=re.IGNORECASE)
    return len(tokens) >= 2


def extract_size_text(body_text: str) -> Optional[str]:
    lines = [line.strip() for line in body_text.splitlines() if line.strip()]

    for i, line in enumerate(lines):
        cleaned = line.replace("×", "x")
        if cleaned.startswith("사이즈"):
            rest = re.sub(r"^사이즈[:\s]*", "", cleaned).strip()
            if looks_like_size_line(rest):
                return rest
            if i + 1 < len(lines) and looks_like_size_line(lines[i + 1]):
                return lines[i + 1].replace("×", "x").strip()

    candidates = []
    for line in lines:
        cleaned = line.replace("×", "x").strip()
        if looks_like_size_line(cleaned):
            candidates.append(cleaned)

    if candidates:
        candidates.sort(
            key=lambda x: len(re.findall(r"(?:W|D|H|SH|AH|Ø)\s*\d+", x, flags=re.IGNORECASE)),
            reverse=True,
        )
        return candidates[0]

    return None


def fetch_html(url: str) -> str:
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://magicfn.com/",
    }
    resp = requests.get(url, headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.text


def extract_image_url_from_html(html: str, base_url: str) -> Optional[str]:
    soup = BeautifulSoup(html, "html.parser")

    selectors = [
        ('meta[property="og:image"]', "content"),
        ('.keyImg img', "src"),
        ('.BigImage', "src"),
        ('.thumbnail img', "src"),
        ('img[alt]', "src"),
    ]

    for selector, attr in selectors:
        for tag in soup.select(selector)[:5]:
            value = tag.get(attr)
            if value and not value.startswith("data:"):
                return normalize_url(value, base_url)

    body_text = soup.get_text("\n", strip=True)
    patterns = [
        r'(//[^"\'<>\s]+/web/product/(?:big|medium|small)/[^"\'<>\s]+\.(?:jpg|jpeg|png|webp))',
        r'(/web/product/(?:big|medium|small)/[^"\'<>\s]+\.(?:jpg|jpeg|png|webp))',
    ]

    for text in [html, body_text]:
        for pattern in patterns:
            m = re.search(pattern, text, flags=re.IGNORECASE)
            if m:
                return normalize_url(m.group(1), base_url)

    return None


def scrape_product(url: str) -> Dict[str, Optional[str]]:
    html = fetch_html(url)
    soup = BeautifulSoup(html, "html.parser")
    body_text = soup.get_text("\n", strip=True)

    product_name = None
    title_meta = soup.select_one('meta[property="og:title"]')
    if title_meta:
        product_name = title_meta.get("content")

    if not product_name:
        h2 = soup.select_one("h2")
        if h2:
            product_name = h2.get_text(" ", strip=True)

    product_name = clean_product_name(product_name)

    vat_price = None
    vat_match = re.search(r"부가세 포함\s*([0-9,]+원)", body_text)
    if vat_match:
        vat_price = clean_price_to_int(vat_match.group(1))

    size_text = extract_size_text(body_text)
    image_url = extract_image_url_from_html(html, url)

    return {
        "url": url,
        "product_name": product_name,
        "image_url": image_url,
        "size_text": size_text,
        "vat_price": vat_price,
        "supply_price": vat_to_supply(vat_price),
    }


def download_image_bytes(image_url: str) -> Optional[bytes]:
    if not image_url:
        return None

    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://magicfn.com/",
    }
    resp = requests.get(image_url, headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.content


def resize_image_for_excel(image_bytes: bytes, max_w: int = 140, max_h: int = 90) -> XLImage:
    pil_img = PILImage.open(BytesIO(image_bytes)).convert("RGB")
    orig_w, orig_h = pil_img.size

    ratio = min(max_w / orig_w, max_h / orig_h)
    new_w = max(1, int(orig_w * ratio))
    new_h = max(1, int(orig_h * ratio))

    pil_img = pil_img.resize((new_w, new_h))

    buf = BytesIO()
    pil_img.save(buf, format="PNG")
    buf.seek(0)

    xl_img = XLImage(buf)
    xl_img.width = new_w
    xl_img.height = new_h
    return xl_img


def snapshot_template(ws):
    max_col = 17
    row_heights = {}
    row_data = {}

    for r in range(15, 28):
        row_heights[r] = ws.row_dimensions[r].height
        row_data[r] = []
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            row_data[r].append({
                "value": cell.value,
                "_style": copy.copy(cell._style),
                "font": copy.copy(cell.font),
                "fill": copy.copy(cell.fill),
                "border": copy.copy(cell.border),
                "alignment": copy.copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy.copy(cell.protection),
            })

    section_merges = [
        mr for mr in ws.merged_cells.ranges
        if mr.min_row >= 15 and mr.max_row <= 27
    ]
    return row_heights, row_data, section_merges


def apply_snapshot_row(ws, row_heights, row_data, src_row: int, dst_row: int):
    max_col = 17
    ws.row_dimensions[dst_row].height = row_heights[src_row]

    for c in range(1, max_col + 1):
        dst = ws.cell(dst_row, c)
        snap = row_data[src_row][c - 1]
        dst.value = snap["value"]
        dst._style = copy.copy(snap["_style"])
        dst.font = copy.copy(snap["font"])
        dst.fill = copy.copy(snap["fill"])
        dst.border = copy.copy(snap["border"])
        dst.alignment = copy.copy(snap["alignment"])
        dst.number_format = snap["number_format"]
        dst.protection = copy.copy(snap["protection"])


def rebuild_quote_section(ws, product_count: int):
    if product_count < 1:
        raise ValueError("상품 개수는 1개 이상이어야 합니다.")

    row_heights, row_data, section_merges = snapshot_template(ws)
    extra_rows = product_count - 1
    section_end = 27 + extra_rows

    # 15행~하단 새 끝행까지의 병합 해제
    to_unmerge = [
        str(mr) for mr in list(ws.merged_cells.ranges)
        if not (mr.max_row < 15 or mr.min_row > section_end)
    ]
    for rng in to_unmerge:
        try:
            ws.unmerge_cells(rng)
        except Exception:
            pass

    # 영역 초기화
    for r in range(15, section_end + 1):
        for c in range(1, 18):
            ws.cell(r, c).value = None

    # 상품행(15행 템플릿 반복)
    product_row_merges = [mr for mr in section_merges if mr.min_row == 15 and mr.max_row == 15]
    for idx in range(product_count):
        row_num = 15 + idx
        apply_snapshot_row(ws, row_heights, row_data, 15, row_num)

        ws[f"B{row_num}"] = None
        ws[f"D{row_num}"] = None
        ws[f"F{row_num}"] = None
        ws[f"L{row_num}"] = None
        ws[f"K{row_num}"] = 1
        ws[f"O{row_num}"] = f"=(L{row_num}-(L{row_num}*0%))"
        ws[f"P{row_num}"] = f"=SUM(O{row_num}*K{row_num})"
        ws[f"Q{row_num}"] = f"=SUM(P{row_num}/10)"

        for mr in product_row_merges:
            ws.merge_cells(
                start_row=row_num,
                start_column=mr.min_col,
                end_row=row_num,
                end_column=mr.max_col,
            )

    # 하단 블록(16~27행) 아래로 이동 복제
    footer_merges = [mr for mr in section_merges if mr.min_row >= 16]
    for src_row in range(16, 28):
        dst_row = src_row + extra_rows
        apply_snapshot_row(ws, row_heights, row_data, src_row, dst_row)

    for mr in footer_merges:
        ws.merge_cells(
            start_row=mr.min_row + extra_rows,
            start_column=mr.min_col,
            end_row=mr.max_row + extra_rows,
            end_column=mr.max_col,
        )

    # 13행 합계 금액 수식 재연결
    total_row = 20 + extra_rows
    ws["E13"] = f"=P{total_row}"
    ws["L13"] = f"=P{total_row}"

    # 운반 설치비 / 하단 합계 수식 재설정
    delivery_row = 16 + extra_rows
    supply_row = 17 + extra_rows
    tax_row = 18 + extra_rows
    cut_row = 19 + extra_rows
    total_row = 20 + extra_rows

    ws[f"K{delivery_row}"] = 1
    ws[f"L{delivery_row}"] = None
    ws[f"O{delivery_row}"] = None
    ws[f"P{delivery_row}"] = f"=SUM(O{delivery_row}*K{delivery_row})"
    ws[f"Q{delivery_row}"] = f"=SUM(P{delivery_row}/10)"

    ws[f"P{supply_row}"] = f"=SUM(P15:P{delivery_row})"
    ws[f"P{tax_row}"] = f"=SUM(Q15:Q{delivery_row})"
    ws[f"P{total_row}"] = f"=SUM(P{supply_row}:Q{cut_row})"


def write_product_row(ws, row_num: int, product: Dict[str, Optional[str]]) -> None:
    ws[f"B{row_num}"] = product["product_name"]
    ws[f"F{row_num}"] = product["size_text"]
    ws[f"K{row_num}"] = 1
    ws[f"L{row_num}"] = product["supply_price"]

    if product["image_url"]:
        image_bytes = download_image_bytes(product["image_url"])
        if image_bytes:
            img = resize_image_for_excel(image_bytes)
            ws.add_image(img, f"D{row_num}")


def build_quote(urls: List[str]) -> tuple[bytes, str]:
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError("앱과 같은 폴더에 매직_견적서_cs.xlsx 파일이 있어야 합니다.")

    wb = load_workbook(TEMPLATE_FILE)
    ws = wb[SHEET_NAME]

    products: List[Dict[str, Optional[str]]] = []
    for idx, url in enumerate(urls, start=1):
        log(f"[{idx}/{len(urls)}] 스크랩 시작")
        log(url)
        product = scrape_product(url)
        products.append(product)
        log(f"[{idx}/{len(urls)}] 스크랩 완료")
        log(f"상품명: {product['product_name']}")
        log(f"이미지: {product['image_url']}")
        log(f"사이즈: {product['size_text']}")
        log(f"공급가: {product['supply_price']}")
        log("-" * 50)

    rebuild_quote_section(ws, len(products))

    for idx, product in enumerate(products):
        row_num = FIRST_ITEM_ROW + idx
        log(f"{row_num}행 입력 시작: {product['product_name']}")
        write_product_row(ws, row_num, product)
        log(f"{row_num}행 입력 완료")

    output_name = f"견적서_자동입력_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_name)
    wb.save(output_path)

    with open(output_path, "rb") as f:
        output_bytes = f.read()

    return output_bytes, output_name


def check_password():
    st.session_state.setdefault("authed", False)

    if st.session_state["authed"]:
        return True

    st.title("매직 견적서 생성기")
    st.caption("비밀번호를 입력해야 사용할 수 있습니다.")

    pw = st.text_input("비밀번호", type="password")
    if st.button("입장"):
        if pw == APP_PASSWORD:
            st.session_state["authed"] = True
            st.rerun()
        else:
            st.error("비밀번호가 틀렸습니다.")
    return False


def main():
    st.set_page_config(page_title="매직 견적서 생성기", layout="wide")

    if not check_password():
        return

    st.title("매직 견적서 생성기")
    st.caption("15행부터 상품이 추가되고, 16~27행 하단 블록은 양식 그대로 아래로 이어집니다.")

    st.session_state.setdefault("logs", [])

    with st.sidebar:
        st.subheader("기준 양식")
        st.write("템플릿 파일: 매직_견적서_cs.xlsx")
        st.write("비밀번호: 0915")
        st.write("필수 파일은 앱 저장소에 함께 있어야 합니다.")

    col1, col2 = st.columns([1.2, 1])

    with col1:
        urls_text = st.text_area(
            "상품 상세페이지 링크",
            height=260,
            placeholder="한 줄에 하나씩 붙여넣으세요.",
        )

        if st.button("견적서 생성", type="primary", use_container_width=True):
            st.session_state["logs"] = []
            urls = [line.strip() for line in urls_text.splitlines() if line.strip()]

            if not urls:
                st.error("상품 링크를 한 줄에 하나씩 입력하세요.")
            else:
                try:
                    output_bytes, output_name = build_quote(urls)
                    st.success("견적서 생성 완료")
                    st.download_button(
                        "결과 파일 다운로드",
                        data=output_bytes,
                        file_name=output_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"오류: {repr(e)}")

    with col2:
        st.subheader("실행 로그")
        st.code("\n".join(st.session_state.get("logs", [])) or "아직 실행 전입니다.", language="text")


if __name__ == "__main__":
    main()
